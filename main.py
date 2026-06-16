import io
import os
import psycopg2
import psycopg2.extras
import telebot
from telebot import types
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from collections import Counter

TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
DATABASE_URL = os.environ.get("DATABASE_URL")
if not TOKEN:
    raise RuntimeError("TELEGRAM_BOT_TOKEN не знайдено в змінних середовища")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL не знайдено в змінних середовища")

bot = telebot.TeleBot(TOKEN)
ADMIN_ID = 1030723047
MEDIA_STATE = "waiting_media"

user_data = {}
admin_state = {}
questions_store = {}
question_counter = 0

# ── DB connection ──────────────────────────────────────────────────────────────

def get_db():
    return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)

# ── DB init ────────────────────────────────────────────────────────────────────

def db_init():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS orders (
                    id SERIAL PRIMARY KEY,
                    client_id BIGINT NOT NULL,
                    client_name TEXT NOT NULL,
                    phone TEXT NOT NULL,
                    filling TEXT NOT NULL,
                    kg TEXT NOT NULL,
                    description TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'new',
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS faq (
                    id SERIAL PRIMARY KEY,
                    question TEXT NOT NULL,
                    answer TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS reviews (
                    id SERIAL PRIMARY KEY,
                    order_id INTEGER NOT NULL,
                    client_id BIGINT NOT NULL,
                    client_name TEXT NOT NULL,
                    rating INTEGER NOT NULL,
                    comment TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS gallery (
                    id SERIAL PRIMARY KEY,
                    file_id TEXT NOT NULL,
                    caption TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS blocked_users (
                    id SERIAL PRIMARY KEY,
                    client_id BIGINT UNIQUE NOT NULL,
                    client_name TEXT,
                    blocked_at TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS templates (
                    id SERIAL PRIMARY KEY,
                    title TEXT NOT NULL,
                    text TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            conn.commit()
            # Seed FAQ if empty
            cur.execute("SELECT COUNT(*) AS cnt FROM faq")
            if cur.fetchone()["cnt"] == 0:
                for q, a in INITIAL_FAQ:
                    cur.execute("INSERT INTO faq (question, answer) VALUES (%s, %s)", (q, a))
                conn.commit()

# ── Orders ─────────────────────────────────────────────────────────────────────

def db_save_order(client_id, name, phone, filling, kg, description):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO orders (client_id, client_name, phone, filling, kg, description, status)
                   VALUES (%s, %s, %s, %s, %s, %s, 'new') RETURNING id""",
                (client_id, name, phone, filling, kg, description)
            )
            row = cur.fetchone()
            conn.commit()
            return row["id"]

def db_get_order(order_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM orders WHERE id = %s", (order_id,))
            return cur.fetchone()

def db_update_status(order_id, status):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE orders SET status = %s WHERE id = %s", (status, order_id))
            conn.commit()

def db_get_all_orders(limit=20):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM orders ORDER BY created_at DESC LIMIT %s", (limit,))
            return cur.fetchall()

def db_delete_order(order_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM orders WHERE id = %s", (order_id,))
            conn.commit()

def db_search_orders(query):
    like = f"%{query}%"
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT * FROM orders
                   WHERE client_name ILIKE %s OR phone ILIKE %s
                   ORDER BY created_at DESC LIMIT 20""",
                (like, like)
            )
            return cur.fetchall()

def db_get_client_orders(client_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM orders WHERE client_id = %s ORDER BY created_at DESC", (client_id,))
            return cur.fetchall()

def db_get_last_client_order(client_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM orders WHERE client_id = %s ORDER BY created_at DESC LIMIT 1",
                (client_id,)
            )
            return cur.fetchone()

def db_get_all_client_ids():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT client_id FROM orders")
            return [row["client_id"] for row in cur.fetchall()]

def db_get_stats():
    with get_db() as conn:
        with conn.cursor() as cur:
            now = datetime.now()
            month_start = now.strftime("%Y-%m-01")

            cur.execute("SELECT COUNT(*) AS cnt FROM orders")
            total = cur.fetchone()["cnt"]

            cur.execute("SELECT COUNT(*) AS cnt FROM orders WHERE created_at >= %s", (month_start,))
            month = cur.fetchone()["cnt"]

            cur.execute("SELECT COUNT(*) AS cnt FROM orders WHERE status = 'done'")
            done = cur.fetchone()["cnt"]

            cur.execute("SELECT COUNT(*) AS cnt FROM orders WHERE status IN ('new','work')")
            active = cur.fetchone()["cnt"]

            cur.execute("SELECT filling FROM orders WHERE status != 'rejected'")
            fillings = [row["filling"] for row in cur.fetchall()]

            cur.execute("SELECT COUNT(DISTINCT client_id) AS cnt FROM orders")
            clients = cur.fetchone()["cnt"]

            return {"total": total, "month": month, "done": done,
                    "active": active, "fillings": fillings, "clients": clients}

# ── Reviews ────────────────────────────────────────────────────────────────────

def db_save_review(order_id, client_id, client_name, rating, comment):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO reviews (order_id, client_id, client_name, rating, comment) VALUES (%s,%s,%s,%s,%s)",
                (order_id, client_id, client_name, rating, comment)
            )
            conn.commit()

def db_get_all_reviews():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM reviews ORDER BY created_at DESC")
            return cur.fetchall()

def db_review_exists(order_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS cnt FROM reviews WHERE order_id = %s", (order_id,))
            return cur.fetchone()["cnt"] > 0

# ── Gallery ────────────────────────────────────────────────────────────────────

def db_add_gallery(file_id, caption):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO gallery (file_id, caption) VALUES (%s,%s) RETURNING id", (file_id, caption))
            row = cur.fetchone()
            conn.commit()
            return row["id"]

def db_get_gallery():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM gallery ORDER BY created_at ASC")
            return cur.fetchall()

def db_delete_gallery(photo_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM gallery WHERE id = %s", (photo_id,))
            conn.commit()

# ── Blocked users ──────────────────────────────────────────────────────────────

def db_block_user(client_id, client_name):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO blocked_users (client_id, client_name)
                   VALUES (%s, %s)
                   ON CONFLICT (client_id) DO UPDATE SET client_name = EXCLUDED.client_name""",
                (client_id, client_name)
            )
            conn.commit()

def db_unblock_user(client_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM blocked_users WHERE client_id = %s", (client_id,))
            conn.commit()

def db_is_blocked(client_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS cnt FROM blocked_users WHERE client_id = %s", (client_id,))
            return cur.fetchone()["cnt"] > 0

def db_get_blocked():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM blocked_users ORDER BY blocked_at DESC")
            return cur.fetchall()

# ── Templates ──────────────────────────────────────────────────────────────────

def db_add_template(title, text):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO templates (title, text) VALUES (%s,%s) RETURNING id", (title, text))
            row = cur.fetchone()
            conn.commit()
            return row["id"]

def db_get_templates():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM templates ORDER BY id ASC")
            return cur.fetchall()

def db_get_template(template_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM templates WHERE id = %s", (template_id,))
            return cur.fetchone()

def db_delete_template(template_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM templates WHERE id = %s", (template_id,))
            conn.commit()

# ── FAQ ────────────────────────────────────────────────────────────────────────

INITIAL_FAQ = [
    ("📦 Яке мінімальне замовлення?",
     "Мінімальне замовлення — *2 кг*.\nЦіна вказана за 1 кг, декор рахується окремо."),
    ("⏱ Скільки часу займає виготовлення?",
     "Зазвичай *2–3 дні*.\nДля святкових тортів рекомендуємо замовляти за *5–7 днів*."),
    ("🚗 Чи є доставка?",
     "Самовивіз — *безкоштовно* біля ТРЦ Вікторія Гарденс.\nДоставка — *за домовленістю*."),
    ("💳 Як оплатити замовлення?",
     "Оплата *при отриманні* готівкою або на картку.\nПередоплата можлива за домовленістю."),
    ("🎨 Чи можна замовити торт за своїм дизайном?",
     "Так! Надішліть фото або опис бажаного дизайну — ми врахуємо всі побажання 🎂"),
    ("🌿 Чи є торти без цукру / для алергіків?",
     "Так, ми можемо адаптувати рецепт. Вкажіть побажання в описі або напишіть нам окремо."),
]

def db_get_faq():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM faq ORDER BY id")
            return cur.fetchall()

def db_get_faq_item(faq_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM faq WHERE id = %s", (faq_id,))
            return cur.fetchone()

def db_add_faq(question, answer):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO faq (question, answer) VALUES (%s,%s) RETURNING id", (question, answer))
            row = cur.fetchone()
            conn.commit()
            return row["id"]

def db_delete_faq(faq_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM faq WHERE id = %s", (faq_id,))
            conn.commit()

# ── Constants ──────────────────────────────────────────────────────────────────

STATUS_LABELS = {
    "new":      "🆕 Нове",
    "work":     "🔧 В роботі",
    "done":     "🎂 Готово",
    "rejected": "❌ Відхилено",
}

CAKE_PRICES = {
    "Бісквіт з фруктами":      1200,
    "Полуничне тірамісу":       1200,
    "Вишня-шоколад":            1300,
    "Лісові ягоди":             1300,
    "Манго-маракуя":            1400,
    "Горіхова карамель-банан":  1400,
    "Орео":                     1400,
    "Фісташка-малина":          1500,
    "Ферреро Роше":             1500,
    "Трюфель":                  1500,
}

CAKES = {
    "Бісквіт з фруктами":      "🍰 Бісквіт + крем + фрукти\n💰 1200 грн/кг",
    "Полуничне тірамісу":       "🍓 Маскарпоне + полуниця\n💰 1200 грн/кг",
    "Вишня-шоколад":            "🍒 Шоколад + вишня\n💰 1300 грн/кг",
    "Лісові ягоди":             "🫐 Ягідний мікс\n💰 1300 грн/кг",
    "Манго-маракуя":            "🥭 Тропічний смак\n💰 1400 грн/кг",
    "Горіхова карамель-банан":  "🍌 Карамель + банан\n💰 1400 грн/кг",
    "Орео":                     "🍪 Крем + Oreo\n💰 1400 грн/кг",
    "Фісташка-малина":          "🌿 Фісташка + малина\n💰 1500 грн/кг",
    "Ферреро Роше":             "🍫 Шоколад + горіх\n💰 1500 грн/кг",
    "Трюфель":                  "🍫 Шоколадний трюфель\n💰 1500 грн/кг",
}

CAKE_PHOTOS = {
    "Бісквіт з фруктами":      "https://images.unsplash.com/photo-1565958011703-44f9829ba187?w=600",
    "Полуничне тірамісу":       "https://images.unsplash.com/photo-1571877227200-a0d98ea607e9?w=600",
    "Вишня-шоколад":            "https://images.unsplash.com/photo-1578985545062-69928b1d9587?w=600",
    "Лісові ягоди":             "https://images.unsplash.com/photo-1488477181946-6428a0291777?w=600",
    "Манго-маракуя":            "https://images.unsplash.com/photo-1519869325930-281384150729?w=600",
    "Горіхова карамель-банан":  "https://images.unsplash.com/photo-1586985289688-ca3cf47d3e6e?w=600",
    "Орео":                     "https://images.unsplash.com/photo-1606890737304-57a1ca8a5b62?w=600",
    "Фісташка-малина":          "https://images.unsplash.com/photo-1562440499-64c9a111f713?w=600",
    "Ферреро Роше":             "https://images.unsplash.com/photo-1599785209707-a456fc1337bb?w=600",
    "Трюфель":                  "https://images.unsplash.com/photo-1611293388250-580b08c4a145?w=600",
}

VICTORIA_GARDENS_LAT = 49.81858
VICTORIA_GARDENS_LON = 23.97621

# ── Helpers ────────────────────────────────────────────────────────────────────

def is_blocked_user(message):
    if message.chat.id == ADMIN_ID:
        return False
    return db_is_blocked(message.chat.id)

def format_order_text(o):
    status_label = STATUS_LABELS.get(o["status"], o["status"])
    created = o["created_at"].strftime("%d.%m.%Y %H:%M") if o["created_at"] else "—"
    return (
        f"📦 *Замовлення #{o['id']}*\n"
        f"📅 {created}\n"
        f"👤 {o['client_name']}\n"
        f"📞 {o['phone']}\n"
        f"🍰 {o['filling']}\n"
        f"⚖️ {o['kg']} кг\n"
        f"✏️ {o['description']}\n"
        f"📊 Статус: {status_label}"
    )

# ── Keyboards ──────────────────────────────────────────────────────────────────

def main_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("🍰 Замовити торт")
    markup.add("📋 Прайс", "📞 Контакти")
    markup.add("❓ Часті питання", "📸 Галерея")
    return markup

def cancel_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("❌ Скасувати")
    return markup

def phone_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton("📱 Поділитися номером", request_contact=True))
    markup.add("❌ Скасувати")
    return markup

def done_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("✅ Готово", "❌ Скасувати")
    return markup

def cakes_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(
        "Бісквіт з фруктами", "Полуничне тірамісу",
        "Вишня-шоколад", "Лісові ягоди",
        "Манго-маракуя", "Горіхова карамель-банан",
        "Орео", "Фісташка-малина",
        "Ферреро Роше", "Трюфель",
        "🔙 Назад"
    )
    return markup

def fillings_inline():
    markup = types.InlineKeyboardMarkup(row_width=1)
    fillings = [
        ("🍰 Бісквіт з фруктами",      "fill_Бісквіт з фруктами"),
        ("🍓 Полуничне тірамісу",       "fill_Полуничне тірамісу"),
        ("🍒 Вишня-шоколад",            "fill_Вишня-шоколад"),
        ("🫐 Лісові ягоди",             "fill_Лісові ягоди"),
        ("🥭 Манго-маракуя",            "fill_Манго-маракуя"),
        ("🍌 Горіхова карамель-банан",  "fill_Горіхова карамель-банан"),
        ("🍪 Орео",                     "fill_Орео"),
        ("🌿 Фісташка-малина",          "fill_Фісташка-малина"),
        ("🍫 Ферреро Роше",             "fill_Ферреро Роше"),
        ("🍫 Трюфель",                  "fill_Трюфель"),
    ]
    for label, data in fillings:
        markup.add(types.InlineKeyboardButton(label, callback_data=data))
    markup.add(types.InlineKeyboardButton("❌ Скасувати замовлення", callback_data="order_cancel"))
    return markup

def faq_keyboard():
    markup = types.InlineKeyboardMarkup()
    for row in db_get_faq():
        markup.add(types.InlineKeyboardButton(row["question"], callback_data=f"faqdb_{row['id']}"))
    markup.add(types.InlineKeyboardButton("✍️ Задати своє питання", callback_data="ask_question"))
    return markup

def build_admin_markup(oid):
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("🔧 Прийнято в роботу", callback_data=f"st_work_{oid}"),
        types.InlineKeyboardButton("🎂 Готово до видачі",  callback_data=f"st_done_{oid}"),
    )
    markup.add(
        types.InlineKeyboardButton("❌ Відхилити",          callback_data=f"st_reject_{oid}"),
        types.InlineKeyboardButton("✉️ Написати клієнту",  callback_data=f"write_{oid}"),
    )
    markup.add(types.InlineKeyboardButton("🗑️ Видалити замовлення", callback_data=f"del_{oid}"))
    return markup

def build_delete_confirm_markup(oid):
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("✅ Так, видалити", callback_data=f"del_yes_{oid}"),
        types.InlineKeyboardButton("❌ Скасувати",     callback_data=f"del_no_{oid}"),
    )
    return markup

def rating_keyboard(order_id):
    markup = types.InlineKeyboardMarkup(row_width=5)
    markup.add(
        types.InlineKeyboardButton("1 ⭐", callback_data=f"rate_{order_id}_1"),
        types.InlineKeyboardButton("2 ⭐", callback_data=f"rate_{order_id}_2"),
        types.InlineKeyboardButton("3 ⭐", callback_data=f"rate_{order_id}_3"),
        types.InlineKeyboardButton("4 ⭐", callback_data=f"rate_{order_id}_4"),
        types.InlineKeyboardButton("5 ⭐", callback_data=f"rate_{order_id}_5"),
    )
    return markup

def gallery_nav_markup(index, total):
    markup = types.InlineKeyboardMarkup(row_width=3)
    buttons = []
    if index > 0:
        buttons.append(types.InlineKeyboardButton("⬅️", callback_data=f"gal_{index - 1}"))
    buttons.append(types.InlineKeyboardButton(f"{index + 1}/{total}", callback_data="gal_noop"))
    if index < total - 1:
        buttons.append(types.InlineKeyboardButton("➡️", callback_data=f"gal_{index + 1}"))
    markup.add(*buttons)
    return markup

def templates_keyboard():
    markup = types.InlineKeyboardMarkup()
    for t in db_get_templates():
        markup.add(types.InlineKeyboardButton(t["title"], callback_data=f"tpl_{t['id']}"))
    markup.add(types.InlineKeyboardButton("➕ Новий шаблон", callback_data="tpl_new"))
    return markup

# ── /start ─────────────────────────────────────────────────────────────────────

@bot.message_handler(commands=['start'])
def start(message):
    if is_blocked_user(message):
        return
    text = (
        "🎂 *Вітаю! Ласкаво просимо до кондитерської!*\n\n"
        "Я допоможу вам замовити торт ручної роботи на будь-яке свято 🎉\n\n"
        "✨ *Що я вмію:*\n"
        "🍰 Прийняти замовлення на торт\n"
        "📋 Показати прайс та начинки\n"
        "📸 Показати галерею наших робіт\n"
        "📞 Дати контакти для зв'язку\n\n"
        "🎯 *Як замовити:*\n"
        "1️⃣ Натисніть *«🍰 Замовити торт»*\n"
        "2️⃣ Вкажіть ім'я та телефон\n"
        "3️⃣ Оберіть начинку та кількість кг\n"
        "4️⃣ Опишіть дизайн і надішліть фото/відео\n"
        "5️⃣ Ми зв'яжемось з вами!\n\n"
        "📌 *Мінімальне замовлення — 2 кг*\n\n"
        "Оберіть потрібний розділ 👇"
    )
    bot.send_message(message.chat.id, text, parse_mode="Markdown", reply_markup=main_keyboard())

# ── /help (admin) ──────────────────────────────────────────────────────────────

@bot.message_handler(commands=['help'])
def admin_help(message):
    if message.chat.id != ADMIN_ID:
        return
    text = (
        "⚙️ *Команди адміністратора*\n\n"
        "📦 *Замовлення:*\n"
        "/orders — останні 20 замовлень\n"
        "/search — пошук за ім'ям або телефоном\n"
        "/export — вивантажити всі замовлення в Excel\n\n"
        "📊 *Статистика:*\n"
        "/stats — замовлення, топ начинок, виторг\n\n"
        "⭐ *Відгуки:*\n"
        "/reviews — всі відгуки клієнтів\n\n"
        "❓ *FAQ:*\n"
        "/addfaq — додати питання до FAQ\n"
        "/delfaq — видалити питання з FAQ\n"
        "/questions — непрочитані питання\n\n"
        "📸 *Галерея:*\n"
        "/addphoto — додати фото\n"
        "/delphoto — видалити фото\n\n"
        "💬 *Шаблонні відповіді:*\n"
        "/templates — надіслати шаблон клієнту\n"
        "/addtemplate — додати шаблон\n"
        "/deltemplate — видалити шаблон\n\n"
        "🚫 *Блокування:*\n"
        "/block [id] — заблокувати клієнта\n"
        "/unblock — розблокувати клієнта\n"
        "/blocked — список заблокованих\n\n"
        "📢 *Розсилка:*\n"
        "/broadcast — повідомлення всім клієнтам\n\n"
        "/help — ця довідка"
    )
    bot.send_message(message.chat.id, text, parse_mode="Markdown")

# ── /stats (admin) ─────────────────────────────────────────────────────────────

@bot.message_handler(commands=['stats'])
def admin_stats(message):
    if message.chat.id != ADMIN_ID:
        return
    s = db_get_stats()
    filling_counter = Counter(s["fillings"])
    top3 = filling_counter.most_common(3)
    top_lines = "\n".join(
        f"  {i+1}. {name} — {cnt} шт." for i, (name, cnt) in enumerate(top3)
    ) if top3 else "  —"
    done_pct = round(s["done"] / s["total"] * 100) if s["total"] else 0
    revenue_min = sum(CAKE_PRICES.get(name, 1300) * cnt * 2 for name, cnt in filling_counter.items())
    text = (
        f"📊 *Статистика кондитерської*\n\n"
        f"📦 Всього замовлень: *{s['total']}*\n"
        f"📅 Цього місяця: *{s['month']}*\n"
        f"✅ Виконано: *{s['done']}* ({done_pct}%)\n"
        f"🔧 Активних: *{s['active']}*\n"
        f"👥 Унікальних клієнтів: *{s['clients']}*\n\n"
        f"🏆 *Топ-3 начинки:*\n{top_lines}\n\n"
        f"💰 *Приблизний виторг* _(мін. 2 кг/замовлення)_:\n"
        f"  ~{revenue_min:,} грн".replace(",", " ")
    )
    bot.send_message(message.chat.id, text, parse_mode="Markdown")

# ── /mystatus (client) ─────────────────────────────────────────────────────────

@bot.message_handler(commands=['mystatus'])
def client_mystatus(message):
    if is_blocked_user(message):
        return
    orders = db_get_client_orders(message.chat.id)
    if not orders:
        bot.send_message(message.chat.id,
            "📭 *У вас ще немає замовлень*\n\nНатисніть «🍰 Замовити торт» щоб зробити перше! 🎂",
            parse_mode="Markdown", reply_markup=main_keyboard())
        return
    active = [o for o in orders if o["status"] in ("new", "work")]
    finished = [o for o in orders if o["status"] in ("done", "rejected")]
    bot.send_message(message.chat.id,
        f"📦 *Ваші замовлення*\n\nАктивних: {len(active)} | Завершених: {len(finished)}",
        parse_mode="Markdown")
    for o in orders[:5]:
        status_label = STATUS_LABELS.get(o["status"], o["status"])
        date_str = o["created_at"].strftime("%d.%m.%Y") if o["created_at"] else "—"
        bot.send_message(message.chat.id,
            f"📦 *Замовлення #{o['id']}*\n📅 {date_str}\n🍰 {o['filling']} — {o['kg']} кг\n📊 {status_label}",
            parse_mode="Markdown")

# ── /repeat (client) ───────────────────────────────────────────────────────────

@bot.message_handler(commands=['repeat'])
def client_repeat(message):
    if is_blocked_user(message):
        return
    last = db_get_last_client_order(message.chat.id)
    if not last:
        bot.send_message(message.chat.id,
            "📭 У вас ще немає замовлень для повторення.\n\nНатисніть «🍰 Замовити торт» 🎂",
            reply_markup=main_keyboard())
        return
    date_str = last["created_at"].strftime("%d.%m.%Y") if last["created_at"] else "—"
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("✅ Так, повторити", callback_data=f"repeat_yes_{last['id']}"),
        types.InlineKeyboardButton("❌ Скасувати",      callback_data="repeat_no"),
    )
    bot.send_message(message.chat.id,
        f"🔁 *Повторити замовлення?*\n\n📅 Попереднє від {date_str}:\n\n"
        f"🍰 Начинка: *{last['filling']}*\n⚖️ Кількість: *{last['kg']} кг*\n✏️ Дизайн: {last['description']}\n\n"
        "Натисніть «Так» — ми створимо нове замовлення з тими ж параметрами.",
        parse_mode="Markdown", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("repeat_"))
def handle_repeat(call):
    if call.data == "repeat_no":
        bot.answer_callback_query(call.id)
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
        return
    if call.data.startswith("repeat_yes_"):
        orig_id = int(call.data.split("_")[2])
        orig = db_get_order(orig_id)
        if not orig:
            bot.answer_callback_query(call.id, "Замовлення не знайдено")
            return
        chat_id = call.message.chat.id
        oid = db_save_order(chat_id, orig["client_name"], orig["phone"],
                            orig["filling"], orig["kg"], orig["description"])
        bot.send_message(ADMIN_ID,
            f"📦 НОВЕ ЗАМОВЛЕННЯ _(повтор)_ #{oid}:\n\n"
            f"👤 {orig['client_name']} (id: {chat_id})\n📞 {orig['phone']}\n"
            f"🍰 {orig['filling']}\n⚖️ {orig['kg']} кг\n✏️ {orig['description']}\n📎 Медіа: немає",
            parse_mode="Markdown", reply_markup=build_admin_markup(oid))
        bot.answer_callback_query(call.id, "✅ Замовлення створено!")
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
        bot.send_message(chat_id,
            f"✅ *Замовлення #{oid} створено!*\n\nМи зв'яжемось з вами найближчим часом 😊",
            parse_mode="Markdown", reply_markup=main_keyboard())

# ── Admin: orders ──────────────────────────────────────────────────────────────

@bot.message_handler(commands=['orders'])
def admin_orders(message):
    if message.chat.id != ADMIN_ID:
        return
    orders = db_get_all_orders()
    if not orders:
        bot.send_message(message.chat.id, "📭 Замовлень ще немає.")
        return
    for o in orders:
        markup = build_admin_markup(o["id"]) if o["status"] not in ("done", "rejected") else None
        bot.send_message(message.chat.id, format_order_text(o), parse_mode="Markdown", reply_markup=markup)

# ── Admin: search ──────────────────────────────────────────────────────────────

@bot.message_handler(commands=['search'])
def admin_search(message):
    if message.chat.id != ADMIN_ID:
        return
    admin_state[message.chat.id] = {"state": "searching"}
    bot.send_message(message.chat.id,
        "🔍 *Пошук замовлень*\n\nВведіть ім'я або номер телефону клієнта:",
        parse_mode="Markdown")

# ── Admin: reviews ─────────────────────────────────────────────────────────────

@bot.message_handler(commands=['reviews'])
def admin_reviews(message):
    if message.chat.id != ADMIN_ID:
        return
    reviews = db_get_all_reviews()
    if not reviews:
        bot.send_message(message.chat.id, "📭 Відгуків ще немає.")
        return
    avg = sum(r["rating"] for r in reviews) / len(reviews)
    bot.send_message(message.chat.id,
        f"⭐ *Відгуки клієнтів*\n\nВсього: {len(reviews)} | Середня оцінка: {avg:.1f} ⭐",
        parse_mode="Markdown")
    for r in reviews:
        date_str = r["created_at"].strftime("%d.%m.%Y") if r["created_at"] else "—"
        stars = "⭐" * r["rating"]
        comment = f"\n💬 _{r['comment']}_" if r.get("comment") else ""
        bot.send_message(message.chat.id,
            f"{stars} *{r['client_name']}*  _#{r['order_id']}_\n📅 {date_str}{comment}",
            parse_mode="Markdown")

# ── Admin: block / unblock / blocked ──────────────────────────────────────────

@bot.message_handler(commands=['block'])
def admin_block(message):
    if message.chat.id != ADMIN_ID:
        return
    parts = message.text.split()
    if len(parts) < 2:
        bot.send_message(message.chat.id,
            "🚫 *Блокування клієнта*\n\nВикористання: `/block 123456789`",
            parse_mode="Markdown")
        return
    try:
        target_id = int(parts[1])
    except ValueError:
        bot.send_message(message.chat.id, "❌ Невірний ID. Тільки цифри.")
        return
    if target_id == ADMIN_ID:
        bot.send_message(message.chat.id, "❌ Не можна заблокувати адміна.")
        return
    db_block_user(target_id, f"id:{target_id}")
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("🔓 Розблокувати", callback_data=f"unblock_{target_id}"))
    bot.send_message(message.chat.id, f"✅ Клієнт `{target_id}` заблокований.",
        parse_mode="Markdown", reply_markup=markup)

@bot.message_handler(commands=['unblock'])
def admin_unblock(message):
    if message.chat.id != ADMIN_ID:
        return
    blocked = db_get_blocked()
    if not blocked:
        bot.send_message(message.chat.id, "📭 Заблокованих немає.")
        return
    markup = types.InlineKeyboardMarkup()
    for u in blocked:
        label = u.get("client_name") or str(u["client_id"])
        markup.add(types.InlineKeyboardButton(f"🔓 {label}", callback_data=f"unblock_{u['client_id']}"))
    bot.send_message(message.chat.id, "🔓 *Оберіть кого розблокувати:*",
        parse_mode="Markdown", reply_markup=markup)

@bot.message_handler(commands=['blocked'])
def admin_blocked_list(message):
    if message.chat.id != ADMIN_ID:
        return
    blocked = db_get_blocked()
    if not blocked:
        bot.send_message(message.chat.id, "📭 Заблокованих немає.")
        return
    lines = ["🚫 *Заблоковані клієнти:*\n"]
    for u in blocked:
        date_str = u["blocked_at"].strftime("%d.%m.%Y") if u["blocked_at"] else "—"
        lines.append(f"• `{u['client_id']}` — {u.get('client_name','')} _(з {date_str})_")
    bot.send_message(message.chat.id, "\n".join(lines), parse_mode="Markdown")

@bot.callback_query_handler(func=lambda call: call.data.startswith("unblock_"))
def handle_unblock(call):
    if call.message.chat.id != ADMIN_ID:
        return
    target_id = int(call.data.split("_")[1])
    db_unblock_user(target_id)
    bot.answer_callback_query(call.id, "✅ Розблоковано")
    bot.edit_message_text(f"✅ Клієнт `{target_id}` розблокований.",
        chat_id=call.message.chat.id, message_id=call.message.message_id, parse_mode="Markdown")

# ── Admin: templates ───────────────────────────────────────────────────────────

@bot.message_handler(commands=['templates'])
def admin_templates(message):
    if message.chat.id != ADMIN_ID:
        return
    tpls = db_get_templates()
    if not tpls:
        bot.send_message(message.chat.id, "💬 Шаблонів ще немає.\n\nДодайте перший через /addtemplate")
        return
    bot.send_message(message.chat.id, "💬 *Шаблонні відповіді:*\n\nОберіть шаблон:",
        parse_mode="Markdown", reply_markup=templates_keyboard())

@bot.callback_query_handler(func=lambda call: call.data == "tpl_new")
def tpl_new_start(call):
    if call.message.chat.id != ADMIN_ID:
        return
    bot.answer_callback_query(call.id)
    admin_state[call.message.chat.id] = {"state": "adding_tpl_title"}
    bot.send_message(call.message.chat.id,
        "💬 *Новий шаблон*\n\nКрок 1/2 — Введіть *назву* шаблону:",
        parse_mode="Markdown")

@bot.callback_query_handler(func=lambda call: call.data.startswith("tpl_") and call.data != "tpl_new" and not call.data.startswith("tpldel_"))
def handle_template_send(call):
    if call.message.chat.id != ADMIN_ID:
        return
    tpl_id = int(call.data.split("_")[1])
    tpl = db_get_template(tpl_id)
    if not tpl:
        bot.answer_callback_query(call.id, "Шаблон не знайдено")
        return
    bot.answer_callback_query(call.id)
    admin_state[call.message.chat.id] = {
        "state": "sending_template",
        "template_text": tpl["text"],
        "template_title": tpl["title"],
    }
    bot.send_message(call.message.chat.id,
        f"💬 *Шаблон:* {tpl['title']}\n\n_{tpl['text']}_\n\nВведіть *ID клієнта* якому надіслати:",
        parse_mode="Markdown")

@bot.message_handler(commands=['addtemplate'])
def admin_add_template(message):
    if message.chat.id != ADMIN_ID:
        return
    admin_state[message.chat.id] = {"state": "adding_tpl_title"}
    bot.send_message(message.chat.id,
        "💬 *Новий шаблон*\n\nКрок 1/2 — Введіть *назву* шаблону:",
        parse_mode="Markdown")

@bot.message_handler(commands=['deltemplate'])
def admin_del_template(message):
    if message.chat.id != ADMIN_ID:
        return
    tpls = db_get_templates()
    if not tpls:
        bot.send_message(message.chat.id, "📭 Шаблонів немає.")
        return
    markup = types.InlineKeyboardMarkup()
    for t in tpls:
        markup.add(types.InlineKeyboardButton(f"🗑️ {t['title']}", callback_data=f"tpldel_{t['id']}"))
    bot.send_message(message.chat.id, "🗑️ *Оберіть шаблон для видалення:*",
        parse_mode="Markdown", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("tpldel_"))
def handle_template_delete(call):
    if call.message.chat.id != ADMIN_ID:
        return
    tpl_id = int(call.data.split("_")[1])
    tpl = db_get_template(tpl_id)
    if not tpl:
        bot.answer_callback_query(call.id, "Шаблон не знайдено")
        return
    db_delete_template(tpl_id)
    bot.answer_callback_query(call.id, "✅ Видалено")
    bot.edit_message_text(f"🗑️ *Шаблон видалено:* {tpl['title']}",
        chat_id=call.message.chat.id, message_id=call.message.message_id, parse_mode="Markdown")

# ── Admin: broadcast ───────────────────────────────────────────────────────────

@bot.message_handler(commands=['broadcast'])
def admin_broadcast(message):
    if message.chat.id != ADMIN_ID:
        return
    clients = db_get_all_client_ids()
    if not clients:
        bot.send_message(message.chat.id, "📭 Немає клієнтів для розсилки.")
        return
    admin_state[message.chat.id] = {"state": "broadcasting", "clients": clients}
    bot.send_message(message.chat.id,
        f"📢 *Розсилка*\n\nОтримувачів: *{len(clients)}*\n\nВведіть текст повідомлення:",
        parse_mode="Markdown")

# ── Admin: gallery ─────────────────────────────────────────────────────────────

@bot.message_handler(commands=['addphoto'])
def admin_add_photo(message):
    if message.chat.id != ADMIN_ID:
        return
    admin_state[message.chat.id] = {"state": "adding_photo"}
    bot.send_message(message.chat.id, "📸 *Додати фото до галереї*\n\nНадішліть фото:",
        parse_mode="Markdown")

@bot.message_handler(commands=['delphoto'])
def admin_del_photo(message):
    if message.chat.id != ADMIN_ID:
        return
    photos = db_get_gallery()
    if not photos:
        bot.send_message(message.chat.id, "📭 Галерея порожня.")
        return
    for i, p in enumerate(photos):
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🗑️ Видалити", callback_data=f"galdel_{p['id']}"))
        bot.send_photo(message.chat.id, p["file_id"],
            caption=p.get("caption") or f"Фото #{i + 1}", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("galdel_"))
def handle_gallery_delete(call):
    if call.message.chat.id != ADMIN_ID:
        return
    db_delete_gallery(int(call.data.split("_")[1]))
    bot.answer_callback_query(call.id, "✅ Фото видалено")
    bot.delete_message(call.message.chat.id, call.message.message_id)

@bot.message_handler(
    content_types=["photo"],
    func=lambda m: m.chat.id == ADMIN_ID and admin_state.get(m.chat.id, {}).get("state") == "adding_photo"
)
def receive_gallery_photo(message):
    file_id = message.photo[-1].file_id
    admin_state[message.chat.id] = {"state": "adding_photo_caption", "file_id": file_id}
    bot.send_message(message.chat.id,
        "✅ Фото отримано!\n\nВведіть підпис _(або «-» щоб пропустити)_:",
        parse_mode="Markdown")

# ── Gallery for clients ────────────────────────────────────────────────────────

@bot.message_handler(func=lambda m: m.text == "📸 Галерея")
def gallery_menu(message):
    if is_blocked_user(message):
        return
    photos = db_get_gallery()
    if not photos:
        bot.send_message(message.chat.id,
            "📷 *Галерея поки порожня*\n\nСкоро тут з'являться фото наших робіт! 🎂",
            parse_mode="Markdown")
        return
    p = photos[0]
    markup = gallery_nav_markup(0, len(photos))
    bot.send_photo(message.chat.id, p["file_id"],
        caption=p.get("caption") or "Наша робота 🎂", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("gal_") and not call.data.startswith("galdel_"))
def gallery_navigate(call):
    if call.data == "gal_noop":
        bot.answer_callback_query(call.id)
        return
    index = int(call.data.split("_")[1])
    photos = db_get_gallery()
    if not photos or index >= len(photos):
        bot.answer_callback_query(call.id, "Фото не знайдено")
        return
    bot.answer_callback_query(call.id)
    p = photos[index]
    markup = gallery_nav_markup(index, len(photos))
    try:
        bot.edit_message_media(
            media=types.InputMediaPhoto(p["file_id"], caption=p.get("caption") or "Наша робота 🎂"),
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            reply_markup=markup
        )
    except Exception:
        bot.send_photo(call.message.chat.id, p["file_id"],
            caption=p.get("caption") or "Наша робота 🎂", reply_markup=markup)

# ── Review flow ────────────────────────────────────────────────────────────────

@bot.callback_query_handler(func=lambda call: call.data.startswith("rate_"))
def handle_review_rating(call):
    parts = call.data.split("_")
    order_id = int(parts[1])
    rating = int(parts[2])
    if db_review_exists(order_id):
        bot.answer_callback_query(call.id, "Ви вже залишили відгук")
        return
    bot.answer_callback_query(call.id)
    bot.edit_message_text(
        f"Ви поставили: {'⭐' * rating}\n\nДякуємо! Напишіть короткий коментар _(або «-» щоб пропустити)_:",
        chat_id=call.message.chat.id, message_id=call.message.message_id, parse_mode="Markdown")
    user_data[call.message.chat.id] = {
        "state": "review_comment",
        "order_id": order_id,
        "rating": rating,
        "client_name": call.from_user.first_name or "Клієнт",
    }

@bot.message_handler(func=lambda m: user_data.get(m.chat.id, {}).get("state") == "review_comment")
def handle_review_comment(message):
    chat_id = message.chat.id
    data = user_data.pop(chat_id, {})
    comment = None if message.text.strip() == "-" else message.text.strip()
    db_save_review(data["order_id"], chat_id, data["client_name"], data["rating"], comment)
    stars = "⭐" * data["rating"]
    bot.send_message(chat_id,
        f"✅ Дякуємо за відгук! {stars}\n\nВаша думка дуже важлива для нас 🎂",
        reply_markup=main_keyboard())
    comment_text = f"\n💬 _{comment}_" if comment else ""
    bot.send_message(ADMIN_ID,
        f"⭐ *Новий відгук!*\n\n👤 {data['client_name']}\n📦 Замовлення #{data['order_id']}\nОцінка: {stars}{comment_text}",
        parse_mode="Markdown")

# ── Price / Contacts ───────────────────────────────────────────────────────────

@bot.message_handler(func=lambda m: m.text == "📋 Прайс")
def price(message):
    if is_blocked_user(message):
        return
    text = (
        "📋 Вимоги:\n\n"
        "• Мінімальне замовлення — 2 кг\n"
        "• Ціна вказана за 1 кг\n"
        "• Декор рахується окремо\n\n"
        "👇 Оберіть начинку:"
    )
    bot.send_message(message.chat.id, text, reply_markup=cakes_keyboard())

@bot.message_handler(func=lambda m: m.text in CAKES)
def cake_info(message):
    if is_blocked_user(message):
        return
    caption = CAKES[message.text]
    photo = CAKE_PHOTOS.get(message.text)
    if photo:
        bot.send_photo(message.chat.id, photo, caption=caption)
    else:
        bot.send_message(message.chat.id, caption)

@bot.message_handler(func=lambda m: m.text == "🔙 Назад")
def back(message):
    if is_blocked_user(message):
        return
    bot.send_message(message.chat.id, "Головне меню:", reply_markup=main_keyboard())

@bot.message_handler(func=lambda m: m.text == "📞 Контакти")
def contacts(message):
    if is_blocked_user(message):
        return
    text = (
        "📍 *Контакти та самовивіз*\n\n"
        "🏬 *Адреса:*\n"
        "м. Львів, вул. Пасічна, 188а\n"
        "_(біля ТРЦ Вікторія Гарденс)_\n\n"
        "📞 *Телефон:*\n"
        "+380XXXXXXXXX\n\n"
        "💬 *Instagram / Viber / WhatsApp:*\n"
        "@cakebot\n\n"
        "🕐 *Години роботи:*\n"
        "Пн–Пт: 09:00 – 19:00\n"
        "Сб–Нд: 10:00 – 18:00\n\n"
        "📦 Самовивіз — безкоштовно\n"
        "🚗 Доставка — за домовленістю\n\n"
        "👇 Натисніть кнопку нижче щоб відкрити локацію на карті:"
    )
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("📍 Показати на карті", callback_data="send_location"))
    bot.send_message(message.chat.id, text, parse_mode="Markdown", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == "send_location")
def send_location(call):
    bot.answer_callback_query(call.id)
    bot.send_location(call.message.chat.id, VICTORIA_GARDENS_LAT, VICTORIA_GARDENS_LON)

# ── FAQ ────────────────────────────────────────────────────────────────────────

@bot.message_handler(func=lambda m: m.text == "❓ Часті питання")
def faq_menu(message):
    if is_blocked_user(message):
        return
    bot.send_message(message.chat.id,
        "❓ *Часті питання*\n\nОберіть питання або задайте своє 👇",
        parse_mode="Markdown", reply_markup=faq_keyboard())

@bot.callback_query_handler(func=lambda call: call.data.startswith("faqdb_"))
def faq_answer(call):
    faq_id = int(call.data.split("_")[1])
    row = db_get_faq_item(faq_id)
    if not row:
        bot.answer_callback_query(call.id, "Питання не знайдено")
        return
    bot.answer_callback_query(call.id)
    back_markup = types.InlineKeyboardMarkup()
    back_markup.add(types.InlineKeyboardButton("⬅️ Назад до питань", callback_data="faq_back"))
    bot.edit_message_text(
        f"❓ *{row['question']}*\n\n{row['answer']}",
        chat_id=call.message.chat.id, message_id=call.message.message_id,
        parse_mode="Markdown", reply_markup=back_markup)

@bot.callback_query_handler(func=lambda call: call.data == "faq_back")
def faq_back(call):
    bot.answer_callback_query(call.id)
    bot.edit_message_text(
        "❓ *Часті питання*\n\nОберіть питання або задайте своє 👇",
        chat_id=call.message.chat.id, message_id=call.message.message_id,
        parse_mode="Markdown", reply_markup=faq_keyboard())

@bot.callback_query_handler(func=lambda call: call.data == "ask_question")
def ask_question_start(call):
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id,
        "✍️ *Задайте ваше питання:*\n\n_Напишіть повідомлення — ми відповімо якнайшвидше_",
        parse_mode="Markdown", reply_markup=cancel_keyboard())
    bot.register_next_step_handler(call.message, receive_client_question)

def receive_client_question(message):
    if message.text == "❌ Скасувати":
        cancel(message)
        return
    global question_counter
    question_counter += 1
    qid = question_counter
    client_id = message.chat.id
    client_name = message.from_user.first_name or "Клієнт"
    questions_store[qid] = {
        "client_id": client_id,
        "client_name": client_name,
        "question": message.text,
        "answered": False,
    }
    admin_markup = types.InlineKeyboardMarkup()
    admin_markup.add(types.InlineKeyboardButton("✏️ Відповісти", callback_data=f"ans_{qid}"))
    bot.send_message(ADMIN_ID,
        f"❓ *Нове питання #{qid}*\n\n👤 {client_name} (id: `{client_id}`)\n\n📝 {message.text}",
        parse_mode="Markdown", reply_markup=admin_markup)
    bot.send_message(message.chat.id, "✅ Ваше питання надіслано! Ми відповімо вам у цьому чаті.",
        reply_markup=main_keyboard())

@bot.message_handler(commands=['addfaq'])
def admin_add_faq(message):
    if message.chat.id != ADMIN_ID:
        return
    admin_state[message.chat.id] = {"state": "adding_faq_q"}
    bot.send_message(message.chat.id,
        "➕ *Додати питання в FAQ*\n\nКрок 1/2 — Введіть *текст питання*:\n_(наприклад: «💰 Яка вартість доставки?»)_",
        parse_mode="Markdown")

@bot.message_handler(commands=['delfaq'])
def admin_del_faq(message):
    if message.chat.id != ADMIN_ID:
        return
    rows = db_get_faq()
    if not rows:
        bot.send_message(message.chat.id, "📭 FAQ порожній.")
        return
    bot.send_message(message.chat.id, "🗑️ *Оберіть питання для видалення:*", parse_mode="Markdown")
    for row in rows:
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🗑️ Видалити", callback_data=f"faqdel_{row['id']}"))
        preview = row["answer"][:80] + ("…" if len(row["answer"]) > 80 else "")
        bot.send_message(message.chat.id,
            f"❓ *{row['question']}*\n\n_{preview}_",
            parse_mode="Markdown", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("faqdel_"))
def handle_faq_delete(call):
    faq_id = int(call.data.split("_")[1])
    row = db_get_faq_item(faq_id)
    if not row:
        bot.answer_callback_query(call.id, "Питання не знайдено")
        return
    db_delete_faq(faq_id)
    bot.answer_callback_query(call.id, "✅ Видалено")
    bot.edit_message_text(f"🗑️ *Видалено:* {row['question']}",
        chat_id=call.message.chat.id, message_id=call.message.message_id, parse_mode="Markdown")

@bot.message_handler(commands=['questions'])
def admin_questions(message):
    if message.chat.id != ADMIN_ID:
        return
    unanswered = {qid: q for qid, q in questions_store.items() if not q["answered"]}
    if not unanswered:
        bot.send_message(message.chat.id, "📭 Немає нових питань.")
        return
    bot.send_message(message.chat.id, f"📬 *Непрочитані питання ({len(unanswered)}):*", parse_mode="Markdown")
    for qid, q in unanswered.items():
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("✏️ Відповісти", callback_data=f"ans_{qid}"))
        bot.send_message(message.chat.id,
            f"❓ *Питання #{qid}*\n👤 {q['client_name']}\n\n📝 {q['question']}",
            parse_mode="Markdown", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("ans_"))
def admin_answer_start(call):
    qid = int(call.data.split("_")[1])
    q = questions_store.get(qid)
    if not q:
        bot.answer_callback_query(call.id, "Питання не знайдено")
        return
    admin_state[call.message.chat.id] = {
        "state": "answering",
        "qid": qid,
        "client_id": q["client_id"],
        "client_name": q["client_name"],
    }
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id,
        f"✏️ Введіть відповідь для *{q['client_name']}* (питання #{qid}):\n\n_«{q['question']}»_",
        parse_mode="Markdown")

# ── Cancel / Order flow ────────────────────────────────────────────────────────

@bot.message_handler(func=lambda m: m.text == "❌ Скасувати")
def cancel(message):
    user_data.pop(message.chat.id, None)
    bot.send_message(message.chat.id, "❌ Замовлення скасовано", reply_markup=main_keyboard())

@bot.message_handler(func=lambda m: m.text == "🍰 Замовити торт")
def order_start(message):
    if is_blocked_user(message):
        return
    bot.send_message(message.chat.id, "Як вас звати?", reply_markup=cancel_keyboard())
    bot.register_next_step_handler(message, get_name)

def get_name(message):
    if message.text == "❌ Скасувати":
        cancel(message)
        return
    user_data[message.chat.id] = {"name": message.text}
    bot.send_message(message.chat.id,
        "Введіть номер телефону або натисніть кнопку нижче 👇",
        reply_markup=phone_keyboard())
    bot.register_next_step_handler(message, get_phone)

def get_phone(message):
    if message.text == "❌ Скасувати":
        cancel(message)
        return
    phone = message.contact.phone_number if message.content_type == "contact" else message.text
    user_data[message.chat.id]["phone"] = phone
    bot.send_message(message.chat.id, "🎂 *Оберіть начинку торту:*",
        parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
    bot.send_message(message.chat.id, "👇", reply_markup=fillings_inline())

@bot.callback_query_handler(func=lambda call: call.data.startswith("fill_"))
def get_filling(call):
    chat_id = call.message.chat.id
    filling = call.data.replace("fill_", "")
    if chat_id not in user_data:
        bot.answer_callback_query(call.id)
        bot.send_message(chat_id, "Сесія закінчилась. Почніть замовлення знову.", reply_markup=main_keyboard())
        return
    user_data[chat_id]["filling"] = filling
    bot.answer_callback_query(call.id, f"✅ Обрано: {filling}")
    bot.edit_message_text(f"🎂 Начинка: *{filling}*",
        chat_id=chat_id, message_id=call.message.message_id, parse_mode="Markdown")
    bot.send_message(chat_id,
        "⚖️ Введіть бажану кількість кілограмів:\n_(мінімум 2 кг)_",
        parse_mode="Markdown", reply_markup=cancel_keyboard())
    bot.register_next_step_handler(call.message, get_kg)

def get_kg(message):
    if message.text == "❌ Скасувати":
        cancel(message)
        return
    user_data[message.chat.id]["kg"] = message.text
    bot.send_message(message.chat.id,
        "✏️ *Опишіть бажаний дизайн торту:*\n\n"
        "_Наприклад: квіти з крему, напис «З Днем Народження», кольорова глазур тощо_",
        parse_mode="Markdown", reply_markup=cancel_keyboard())
    bot.register_next_step_handler(message, get_description)

def get_description(message):
    if message.text == "❌ Скасувати":
        cancel(message)
        return
    chat_id = message.chat.id
    user_data[chat_id]["description"] = message.text
    user_data[chat_id]["state"] = MEDIA_STATE
    user_data[chat_id]["media"] = []
    bot.send_message(chat_id,
        "📸 *Надішліть фото або відео як приклад дизайну*\n\n"
        "Можна надіслати скільки завгодно — по одному або альбомом 📂\n\n"
        "Коли закінчите — натисніть *✅ Готово*",
        parse_mode="Markdown", reply_markup=done_keyboard())

@bot.message_handler(
    content_types=["photo", "video"],
    func=lambda m: user_data.get(m.chat.id, {}).get("state") == MEDIA_STATE
)
def collect_media(message):
    chat_id = message.chat.id
    if message.content_type == "photo":
        user_data[chat_id]["media"].append({"type": "photo", "file_id": message.photo[-1].file_id})
        bot.send_message(chat_id, f"✅ Фото додано (всього: {len(user_data[chat_id]['media'])})")
    elif message.content_type == "video":
        user_data[chat_id]["media"].append({"type": "video", "file_id": message.video.file_id})
        bot.send_message(chat_id, f"✅ Відео додано (всього: {len(user_data[chat_id]['media'])})")

@bot.message_handler(
    func=lambda m: m.text == "✅ Готово" and user_data.get(m.chat.id, {}).get("state") == MEDIA_STATE
)
def finish_order(message):
    chat_id = message.chat.id
    data = user_data.get(chat_id, {})
    media_list = data.get("media", [])
    summary = (
        "📋 *Перевірте ваше замовлення:*\n\n"
        f"👤 Ім'я: {data.get('name', '—')}\n"
        f"📞 Телефон: {data.get('phone', '—')}\n"
        f"🍰 Начинка: {data.get('filling', '—')}\n"
        f"⚖️ Кількість: {data.get('kg', '—')} кг\n"
        f"✏️ Дизайн: {data.get('description', '—')}\n"
        f"📎 Фото/відео: {len(media_list) if media_list else 'немає'}\n\n"
        "Все вірно?"
    )
    confirm_markup = types.InlineKeyboardMarkup()
    confirm_markup.add(
        types.InlineKeyboardButton("✅ Підтвердити", callback_data="order_confirm"),
        types.InlineKeyboardButton("❌ Скасувати",   callback_data="order_cancel"),
    )
    bot.send_message(chat_id, "⏳", reply_markup=types.ReplyKeyboardRemove())
    bot.send_message(chat_id, summary, parse_mode="Markdown", reply_markup=confirm_markup)

@bot.callback_query_handler(func=lambda call: call.data in ("order_confirm", "order_cancel"))
def handle_order_confirm(call):
    chat_id = call.message.chat.id
    if call.data == "order_cancel":
        user_data.pop(chat_id, None)
        bot.answer_callback_query(call.id)
        bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=None)
        bot.send_message(chat_id, "❌ Замовлення скасовано", reply_markup=main_keyboard())
        return
    data = user_data.get(chat_id, {})
    media_list = data.get("media", [])
    oid = db_save_order(
        client_id=chat_id,
        name=data.get("name", "—"),
        phone=data.get("phone", "—"),
        filling=data.get("filling", "—"),
        kg=data.get("kg", "—"),
        description=data.get("description", "—"),
    )
    bot.send_message(ADMIN_ID,
        f"📦 НОВЕ ЗАМОВЛЕННЯ #{oid}:\n\n"
        f"👤 {data.get('name','—')} (id: {chat_id})\n"
        f"📞 {data.get('phone','—')}\n"
        f"🍰 {data.get('filling','—')}\n"
        f"⚖️ {data.get('kg','—')} кг\n"
        f"✏️ {data.get('description','—')}\n"
        f"📎 Медіа: {len(media_list)} шт.",
        parse_mode="Markdown", reply_markup=build_admin_markup(oid))
    for item in media_list:
        if item["type"] == "photo":
            bot.send_photo(ADMIN_ID, item["file_id"])
        elif item["type"] == "video":
            bot.send_video(ADMIN_ID, item["file_id"])
    user_data.pop(chat_id, None)
    bot.answer_callback_query(call.id, "✅ Замовлення відправлено!")
    bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=None)
    bot.send_message(chat_id,
        f"✅ Дякуємо! Ваше замовлення №{oid} прийнято.\nМи скоро з вами зв'яжемось 😊",
        reply_markup=main_keyboard())

# ── Status handler ─────────────────────────────────────────────────────────────

@bot.callback_query_handler(func=lambda call: call.data.startswith("st_"))
def handle_status(call):
    parts = call.data.split("_")
    status = parts[1]
    oid = int(parts[2])
    order = db_get_order(oid)
    if not order:
        bot.answer_callback_query(call.id, "Замовлення не знайдено")
        return
    client_id = order["client_id"]
    name = order["client_name"]
    if status == "work":
        client_msg = (f"🔧 *Замовлення №{oid} прийнято в роботу!*\n\n"
                      f"Привіт, {name}! Ваш торт вже готується 🎂\n"
                      "Ми повідомимо вас коли він буде готовий.")
        admin_confirm = f"✅ Статус №{oid} → В роботі"
        db_update_status(oid, "work")
    elif status == "reject":
        client_msg = (f"❌ *Замовлення №{oid} відхилено.*\n\n"
                      f"Привіт, {name}! На жаль, ми не можемо виконати це замовлення.\nЗв'яжіться з нами 📞")
        admin_confirm = f"❌ Замовлення №{oid} відхилено"
        db_update_status(oid, "rejected")
    else:
        client_msg = (f"🎂 *Замовлення №{oid} готове!*\n\n"
                      f"Привіт, {name}! Ваш торт готовий до видачі 🎉\n"
                      "Зв'яжіться з нами для уточнення деталей.")
        admin_confirm = f"✅ Статус №{oid} → Готово до видачі"
        db_update_status(oid, "done")
    try:
        bot.send_message(client_id, client_msg, parse_mode="Markdown")
        if status == "done":
            bot.send_message(client_id,
                "⭐ *Будь ласка, залиште відгук!*\n\nОцініть якість від 1 до 5:",
                parse_mode="Markdown", reply_markup=rating_keyboard(oid))
    except Exception:
        pass
    updated_order = db_get_order(oid)
    bot.answer_callback_query(call.id, admin_confirm)
    bot.edit_message_text(format_order_text(updated_order),
        chat_id=call.message.chat.id, message_id=call.message.message_id,
        parse_mode="Markdown", reply_markup=None)

@bot.callback_query_handler(func=lambda call: call.data.startswith("write_"))
def handle_write_to_client(call):
    oid = int(call.data.split("_")[1])
    order = db_get_order(oid)
    if not order:
        bot.answer_callback_query(call.id, "Замовлення не знайдено")
        return
    admin_state[call.message.chat.id] = {
        "state": "writing",
        "order_id": oid,
        "client_id": order["client_id"],
        "client_name": order["client_name"],
    }
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id,
        f"✏️ Введіть повідомлення для *{order['client_name']}* (замовлення №{oid}):\n\n"
        "_Надішліть текст — він буде доставлений клієнту від імені бота_",
        parse_mode="Markdown")

# ── Admin state machine ────────────────────────────────────────────────────────

ADMIN_STATES = (
    "writing", "answering", "adding_faq_q", "adding_faq_a",
    "broadcasting", "searching", "adding_photo_caption",
    "adding_tpl_title", "adding_tpl_text", "sending_template",
)

@bot.message_handler(
    func=lambda m: m.chat.id == ADMIN_ID and admin_state.get(m.chat.id, {}).get("state") in ADMIN_STATES
)
def admin_reply_handler(message):
    state = admin_state.get(message.chat.id, {})
    current = state.get("state")

    if current == "writing":
        admin_state.pop(message.chat.id, None)
        bot.send_message(state["client_id"],
            f"📩 *Повідомлення від кондитерської*\n\n{message.text}", parse_mode="Markdown")
        bot.send_message(message.chat.id,
            f"✅ Повідомлення надіслано {state.get('client_name')} (замовлення №{state.get('order_id')})")

    elif current == "answering":
        admin_state.pop(message.chat.id, None)
        qid = state.get("qid")
        q = questions_store.get(qid, {})
        bot.send_message(state["client_id"],
            f"💬 *Відповідь на ваше питання:*\n\n_«{q.get('question','')}»_\n\n{message.text}",
            parse_mode="Markdown")
        if qid in questions_store:
            questions_store[qid]["answered"] = True
        bot.send_message(message.chat.id,
            f"✅ Відповідь надіслано {state.get('client_name')} (питання #{qid})")

    elif current == "adding_faq_q":
        admin_state[message.chat.id] = {"state": "adding_faq_a", "question": message.text}
        bot.send_message(message.chat.id,
            f"✅ Питання збережено:\n_«{message.text}»_\n\nКрок 2/2 — Введіть *відповідь* на це питання:",
            parse_mode="Markdown")

    elif current == "adding_faq_a":
        admin_state.pop(message.chat.id, None)
        faq_id = db_add_faq(state.get("question", ""), message.text)
        bot.send_message(message.chat.id,
            f"✅ *Питання #{faq_id} додано до FAQ!*\n\n❓ {state.get('question')}\n\n💬 {message.text}",
            parse_mode="Markdown")

    elif current == "searching":
        admin_state.pop(message.chat.id, None)
        results = db_search_orders(message.text.strip())
        if not results:
            bot.send_message(message.chat.id,
                f"🔍 За запитом *«{message.text}»* нічого не знайдено.", parse_mode="Markdown")
            return
        bot.send_message(message.chat.id, f"🔍 Знайдено: *{len(results)}* замовлень", parse_mode="Markdown")
        for o in results:
            markup = build_admin_markup(o["id"]) if o["status"] not in ("done", "rejected") else None
            bot.send_message(message.chat.id, format_order_text(o), parse_mode="Markdown", reply_markup=markup)

    elif current == "broadcasting":
        admin_state.pop(message.chat.id, None)
        clients = state.get("clients", [])
        sent = failed = 0
        for cid in clients:
            try:
                bot.send_message(cid, f"✨ *Маємо для вас інформацію:*\n\n{message.text}",
                    parse_mode="Markdown")
                sent += 1
            except Exception:
                failed += 1
        bot.send_message(message.chat.id,
            f"✅ Розсилку завершено!\n\n📤 Надіслано: {sent}\n❌ Не доставлено: {failed}")

    elif current == "adding_photo_caption":
        admin_state.pop(message.chat.id, None)
        caption = None if message.text.strip() == "-" else message.text.strip()
        photo_id = db_add_gallery(state["file_id"], caption)
        bot.send_message(message.chat.id,
            f"✅ *Фото #{photo_id} додано до галереї!*\nПідпис: _{caption or 'без підпису'}_",
            parse_mode="Markdown")

    elif current == "adding_tpl_title":
        admin_state[message.chat.id] = {"state": "adding_tpl_text", "title": message.text}
        bot.send_message(message.chat.id,
            f"✅ Назва: *{message.text}*\n\nКрок 2/2 — Введіть *текст* шаблону:",
            parse_mode="Markdown")

    elif current == "adding_tpl_text":
        admin_state.pop(message.chat.id, None)
        tpl_id = db_add_template(state["title"], message.text)
        bot.send_message(message.chat.id,
            f"✅ *Шаблон #{tpl_id} збережено!*\n\n📌 {state['title']}\n\n_{message.text}_",
            parse_mode="Markdown")

    elif current == "sending_template":
        admin_state.pop(message.chat.id, None)
        try:
            target_id = int(message.text.strip())
            bot.send_message(target_id,
                f"📩 *Повідомлення від кондитерської*\n\n{state['template_text']}",
                parse_mode="Markdown")
            bot.send_message(message.chat.id,
                f"✅ Шаблон «{state['template_title']}» надіслано клієнту `{target_id}`",
                parse_mode="Markdown")
        except Exception as e:
            bot.send_message(message.chat.id,
                f"❌ Не вдалося надіслати. Перевірте ID.\n_{e}_", parse_mode="Markdown")

# ── Delete order ───────────────────────────────────────────────────────────────

@bot.callback_query_handler(func=lambda call: call.data.startswith("del_"))
def handle_delete(call):
    parts = call.data.split("_", 2)
    if len(parts) == 2:
        oid = int(parts[1])
        if not db_get_order(oid):
            bot.answer_callback_query(call.id, "Замовлення не знайдено")
            return
        bot.answer_callback_query(call.id)
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id,
            reply_markup=build_delete_confirm_markup(oid))
    else:
        action = parts[1]
        oid = int(parts[2])
        if action == "yes":
            db_delete_order(oid)
            bot.answer_callback_query(call.id, f"🗑️ Замовлення №{oid} видалено")
            bot.edit_message_text(f"🗑️ *Замовлення #{oid} видалено*",
                chat_id=call.message.chat.id, message_id=call.message.message_id, parse_mode="Markdown")
        elif action == "no":
            order = db_get_order(oid)
            bot.answer_callback_query(call.id, "Скасовано")
            markup = build_admin_markup(oid) if order else None
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=markup)

# ── Export ─────────────────────────────────────────────────────────────────────

@bot.message_handler(commands=['export'])
def export_orders(message):
    if message.chat.id != ADMIN_ID:
        return
    orders = db_get_all_orders(limit=10000)
    if not orders:
        bot.send_message(message.chat.id, "📭 Замовлень ще немає.")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Замовлення"
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    headers = ["#", "Дата", "Ім'я", "Телефон", "Начинка", "Кг", "Дизайн", "Статус"]
    col_widths = [5, 18, 18, 18, 22, 6, 35, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22
    status_colors = {"new": "FFF9C4", "work": "C8E6C9", "done": "BBDEFB", "rejected": "FFCDD2"}
    for row_idx, o in enumerate(orders, start=2):
        status_key = o["status"]
        status_label = STATUS_LABELS.get(status_key, status_key)
        created = o["created_at"].strftime("%d.%m.%Y %H:%M") if o["created_at"] else "—"
        row_fill = PatternFill(
            start_color=status_colors.get(status_key, "FFFFFF"),
            end_color=status_colors.get(status_key, "FFFFFF"),
            fill_type="solid"
        )
        values = [o["id"], created, o["client_name"], o["phone"],
                  o["filling"], o["kg"], o["description"], status_label]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 7))
        ws.row_dimensions[row_idx].height = 18
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"замовлення_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    bot.send_document(message.chat.id, (filename, buf),
        caption=f"📊 Експорт замовлень — {len(orders)} шт.\n📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}")

# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    db_init()
    print("Бот запущено...")
    bot.polling(none_stop=True)
